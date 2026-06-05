#!/usr/bin/env python3
"""Build and fill the Excel workbook without relying on a bundled .xlsx template."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BLUE = "FF0070C0"
BLACK = "FF000000"
LIGHT_BLUE_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
THIN_GRAY = Side(style="thin", color="FFD9D9D9")
THIN_BLUE = Side(style="thin", color=BLUE)
ARIAL = "Arial"

LABEL_COLUMN = "A"
CHINESE_COLUMN = "B"
HISTORICAL_COLUMNS = ["C", "D", "E", "F", "G"]
FORECAST_COLUMNS = ["H", "I", "J", "K", "L"]
ALL_YEAR_COLUMNS = HISTORICAL_COLUMNS + FORECAST_COLUMNS

ROW_LABELS = {
    1: "RMBm",
    2: "Revenue",
    3: "Growth rate",
    4: "COGS",
    5: "Gross profit",
    6: "Gross margin",
    8: "Business tax and surcharges",
    9: "As % of sales",
    10: "Sales & marketing expenses",
    11: "As % of sales",
    12: "General & administration expenses",
    13: "As % of sales",
    14: "Research expenses",
    15: "As % of sales",
    16: "Other operating expenses",
    17: "As % of sales",
    19: "Operating profit",
    20: "Operating profit margin",
    22: "Net interest expenses",
    23: "As % of sales",
    24: "Other non-operating income / expenses",
    25: "As % of sales",
    26: "Profit before tax",
    27: "Tax expenses",
    28: "Net income",
    29: "Net income margin",
    31: "Assumptions",
    32: "Revenue growth",
    33: "Gross margin",
    34: "Business tax and surcharges as % of sales",
    35: "Sales & marketing expenses as % of sales",
    36: "General & administration expenses as % of sales",
    37: "Research expenses as a % of sales",
    38: "Other operating expenses as % of sales",
    39: "Net interest expenses as % of sales",
    40: "Other non-operating income / expenses as % of sales",
    41: "Corporate income tax rate",
}

ROW_LABELS_ZH = {
    1: "",
    2: "营业收入",
    3: "增长率",
    4: "营业成本",
    5: "毛利润",
    6: "毛利率",
    8: "营业税金及附加",
    9: "占收入比例",
    10: "销售费用",
    11: "占收入比例",
    12: "管理费用",
    13: "占收入比例",
    14: "研发费用",
    15: "占收入比例",
    16: "其他营业费用",
    17: "占收入比例",
    19: "营业利润",
    20: "营业利润率",
    22: "净利息费用",
    23: "占收入比例",
    24: "其他营业外收入/支出",
    25: "占收入比例",
    26: "税前利润",
    27: "所得税",
    28: "净利润",
    29: "净利率",
    31: "假设",
    32: "收入增长率",
    33: "毛利率",
    34: "营业税金及附加率",
    35: "销售费用率",
    36: "管理费用率",
    37: "研发费用率",
    38: "其他营业费用率",
    39: "净利息费用率",
    40: "其他营业外损益率",
    41: "所得税率",
}

INPUT_ROWS = {2, 5, 8, 10, 12, 14, 19, 22, 26, 27}
ASSUMPTION_ROWS = {32, 33, 34, 35, 36, 37, 38, 39, 40, 41}


def load_mapping(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_payload(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def set_formula(ws, cell: str, formula: str) -> None:
    ws[cell] = formula
    ws[cell].font = Font(name=ARIAL, color=BLACK)


def build_blank_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws[f"{LABEL_COLUMN}1"] = "RMBm"
    ws[f"{CHINESE_COLUMN}1"] = ""
    headers = ["2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
    for col, header in zip(ALL_YEAR_COLUMNS, headers):
        ws[f"{col}1"] = header
        ws[f"{col}1"].font = Font(name=ARIAL, color=BLACK, bold=True)

    for row, label in ROW_LABELS.items():
        ws[f"{LABEL_COLUMN}{row}"] = label
        ws[f"{LABEL_COLUMN}{row}"].font = Font(name=ARIAL, color=BLACK, bold=(row in {1, 31}))
        ws[f"{LABEL_COLUMN}{row}"].alignment = Alignment(horizontal="left")

    for row, label in ROW_LABELS_ZH.items():
        ws[f"{CHINESE_COLUMN}{row}"] = label
        ws[f"{CHINESE_COLUMN}{row}"].font = Font(name=ARIAL, color=BLACK, bold=(row in {31}))
        ws[f"{CHINESE_COLUMN}{row}"].alignment = Alignment(horizontal="left")

    ws.column_dimensions[LABEL_COLUMN].width = 32
    ws.column_dimensions[CHINESE_COLUMN].width = 22
    for col in ALL_YEAR_COLUMNS:
        ws.column_dimensions[col].width = 12
        for row in range(1, 42):
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
            ws[f"{col}{row}"].border = Border(bottom=THIN_GRAY)
            if ws[f"{col}{row}"].font.name != ARIAL:
                ws[f"{col}{row}"].font = Font(name=ARIAL, color=BLACK)

    for row in INPUT_ROWS:
        for col in HISTORICAL_COLUMNS:
            ws[f"{col}{row}"].font = Font(name=ARIAL, color=BLUE)
    for row in ASSUMPTION_ROWS:
        for col in FORECAST_COLUMNS:
            ws[f"{col}{row}"].font = Font(name=ARIAL, color=BLUE)
            ws[f"{col}{row}"].fill = LIGHT_BLUE_FILL
            ws[f"{col}{row}"].border = Border(left=THIN_BLUE, right=THIN_BLUE, top=THIN_BLUE, bottom=THIN_BLUE)

    amount_rows = [2, 4, 5, 8, 10, 12, 14, 16, 19, 22, 24, 26, 27, 28]
    pct_rows = [3, 6, 9, 11, 13, 15, 17, 20, 23, 25, 29, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41]
    for row in amount_rows:
        for col in ALL_YEAR_COLUMNS:
            ws[f"{col}{row}"].number_format = '#,##0.0_);\\(#,##0.0\\)'
    for row in pct_rows:
        for col in ALL_YEAR_COLUMNS:
            ws[f"{col}{row}"].number_format = "0.0%"

    for current, previous in zip(HISTORICAL_COLUMNS[1:], HISTORICAL_COLUMNS[:-1]):
        set_formula(ws, f"{current}3", f"={current}2/{previous}2-1")
    for col in ALL_YEAR_COLUMNS:
        set_formula(ws, f"{col}4", f"={col}5-{col}2")
        set_formula(ws, f"{col}6", f"={col}5/{col}$2")
        set_formula(ws, f"{col}9", f"={col}8/{col}$2")
        set_formula(ws, f"{col}11", f"={col}10/{col}$2")
        set_formula(ws, f"{col}13", f"={col}12/{col}$2")
        set_formula(ws, f"{col}15", f"={col}14/{col}$2")
        set_formula(ws, f"{col}17", f"={col}16/{col}$2")
        set_formula(ws, f"{col}20", f"={col}19/{col}$2")
        set_formula(ws, f"{col}23", f"={col}22/{col}$2")
        set_formula(ws, f"{col}25", f"={col}24/{col}$2")
        set_formula(ws, f"{col}28", f"={col}26+{col}27")
        set_formula(ws, f"{col}29", f"={col}28/{col}$2")

    for col in HISTORICAL_COLUMNS:
        set_formula(ws, f"{col}16", f"={col}19-{col}14-{col}12-{col}10-{col}5-{col}8")
        set_formula(ws, f"{col}24", f"={col}26-{col}19-{col}22")
        if col != HISTORICAL_COLUMNS[0]:
            set_formula(ws, f"{col}32", f"={col}3")
        set_formula(ws, f"{col}33", f"={col}6")
        set_formula(ws, f"{col}34", f"={col}9")
        set_formula(ws, f"{col}35", f"={col}11")
        set_formula(ws, f"{col}36", f"={col}13")
        set_formula(ws, f"{col}37", f"={col}15")
        set_formula(ws, f"{col}38", f"={col}17")
        set_formula(ws, f"{col}39", f"={col}23")
        set_formula(ws, f"{col}40", f"={col}25")
        set_formula(ws, f"{col}41", f"=-{col}27/{col}26")
    ws[f"{HISTORICAL_COLUMNS[0]}32"] = None
    ws[f"{HISTORICAL_COLUMNS[0]}32"].font = Font(name=ARIAL, color=BLACK)

    for current, previous in zip(FORECAST_COLUMNS, [HISTORICAL_COLUMNS[-1]] + FORECAST_COLUMNS[:-1]):
        set_formula(ws, f"{current}2", f"={previous}2*(1+{current}32)")
        set_formula(ws, f"{current}3", f"={current}32")
        set_formula(ws, f"{current}5", f"={current}$2*{current}33")
        set_formula(ws, f"{current}8", f"={current}$2*{current}34")
        set_formula(ws, f"{current}10", f"={current}$2*{current}35")
        set_formula(ws, f"{current}12", f"={current}$2*{current}36")
        set_formula(ws, f"{current}14", f"={current}$2*{current}37")
        set_formula(ws, f"{current}16", f"={current}$2*{current}38")
        set_formula(ws, f"{current}19", f"={current}5+{current}10+{current}12+{current}14+{current}16+{current}8")
        set_formula(ws, f"{current}22", f"={current}$2*{current}39")
        set_formula(ws, f"{current}24", f"={current}$2*{current}40")
        set_formula(ws, f"{current}26", f"={current}19+{current}22+{current}24")
        set_formula(ws, f"{current}27", f"=IF({current}26<0,0,-{current}26*{current}41)")
        set_formula(ws, f"{current}32", f"={previous}32")
        set_formula(ws, f"{current}33", f"={previous}33")
        set_formula(ws, f"{current}34", f"={previous}34")
        set_formula(ws, f"{current}35", f"={previous}35")
        set_formula(ws, f"{current}36", f"={previous}36")
        set_formula(ws, f"{current}37", f"={previous}37")
        set_formula(ws, f"{current}38", f"={previous}38")
        set_formula(ws, f"{current}39", f"={previous}39")
        set_formula(ws, f"{current}40", f"={previous}40")
        set_formula(ws, f"{current}41", f"={previous}41")

    for row in [1, 19, 26, 28, 31]:
        for col in [LABEL_COLUMN, CHINESE_COLUMN, *ALL_YEAR_COLUMNS]:
            ws[f"{col}{row}"].font = Font(name=ARIAL, color=BLACK, bold=True)

    return wb


def fill_historical_block(ws, mapping: dict, payload: dict) -> None:
    columns = mapping["historical_columns"]
    row_mapping = mapping["row_mapping"]
    financials = payload.get("historical_financials", {})

    for metric_name, row_number in row_mapping.items():
        values = financials.get(metric_name, [])
        for index, column_letter in enumerate(columns):
            value = values[index] if index < len(values) else None
            ws[f"{column_letter}{row_number}"] = value
            ws[f"{column_letter}{row_number}"].font = Font(name=ARIAL, color=BLUE)


def main() -> int:
    if len(sys.argv) != 4:
        print("Usage: fill_template.py <mapping.json> <payload.json> <output.xlsx>", file=sys.stderr)
        return 1

    mapping_path = Path(sys.argv[1])
    payload_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    mapping = load_mapping(mapping_path)
    payload = load_payload(payload_path)

    workbook = build_blank_template()
    worksheet = workbook[mapping["sheet"]]
    fill_historical_block(worksheet, mapping, payload)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
